"""Tibetan Spirit v7 — Multi-tab xlsx workbook builder.

Reads model_v7.yaml, runs the P&L engine, and generates a 7-tab
spreadsheet suitable for Google Sheets import.

Usage:
    python3 scripts/financial_model/build_xlsx.py
    python3 scripts/financial_model/build_xlsx.py --config path/to/config.yaml
    python3 scripts/financial_model/build_xlsx.py --output path/to/output.xlsx
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter

# Resolve project root for imports
_PROJECT_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(_PROJECT_ROOT))

from scripts.financial_model.model import build_monthly_pnl, build_scenarios

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------

_DARK_BG = PatternFill(start_color="1A252F", end_color="1A252F", fill_type="solid")
_SECTION_BG = PatternFill(start_color="EAF2F8", end_color="EAF2F8", fill_type="solid")
_WHITE_FONT = Font(color="FFFFFF", bold=True, size=11)
_BOLD_FONT = Font(bold=True, size=11)
_SECTION_FONT = Font(bold=True, size=11, color="1A252F")
_INDENT_FONT = Font(color="808080", size=10)
_TITLE_FONT = Font(bold=True, size=14, color="1A252F")
_THIN_BORDER = Border(
    bottom=Side(style="thin", color="D0D0D0"),
)
_TOTAL_BORDER = Border(
    top=Side(style="thin", color="1A252F"),
    bottom=Side(style="double", color="1A252F"),
)


def _make_styles(wb: Workbook) -> dict:
    """Register named styles for dollar and pct formatting."""
    styles = {}

    dollar = NamedStyle(name="dollar")
    dollar.number_format = '#,##0;[Red]-#,##0'
    dollar.font = Font(size=10)
    wb.add_named_style(dollar)
    styles["dollar"] = dollar

    dollar_bold = NamedStyle(name="dollar_bold")
    dollar_bold.number_format = '#,##0;[Red]-#,##0'
    dollar_bold.font = Font(bold=True, size=10)
    wb.add_named_style(dollar_bold)
    styles["dollar_bold"] = dollar_bold

    pct = NamedStyle(name="pct")
    pct.number_format = '0.0%'
    pct.font = Font(size=10)
    wb.add_named_style(pct)
    styles["pct"] = pct

    pct_bold = NamedStyle(name="pct_bold")
    pct_bold.number_format = '0.0%'
    pct_bold.font = Font(bold=True, size=10)
    wb.add_named_style(pct_bold)
    styles["pct_bold"] = pct_bold

    return styles


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _set_col_widths(ws, width_map: dict):
    """Set column widths. width_map: {col_letter_or_num: width}."""
    for col, w in width_map.items():
        if isinstance(col, int):
            col = get_column_letter(col)
        ws.column_dimensions[col].width = w


def _write_header_row(ws, row: int, labels: list[str], start_col: int = 1):
    """Write a dark header row across columns."""
    for i, label in enumerate(labels):
        cell = ws.cell(row=row, column=start_col + i, value=label)
        cell.fill = _DARK_BG
        cell.font = _WHITE_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _write_section_header(ws, row: int, label: str, col_span: int = 1):
    """Write a light-blue section header."""
    cell = ws.cell(row=row, column=1, value=label)
    cell.fill = _SECTION_BG
    cell.font = _SECTION_FONT
    for c in range(2, col_span + 1):
        ws.cell(row=row, column=c).fill = _SECTION_BG


def _write_monthly_row(
    ws, row: int, label: str, values: list[float],
    style_name: str = "dollar", indent: bool = False,
    bold: bool = False, yearly_totals: list[float] | None = None,
    border=None,
):
    """Write a row with label in col A, 36 monthly values in B-AK,
    and optional Y1/Y2/Y3 totals in AL/AM/AN (cols 38-40)."""
    cell = ws.cell(row=row, column=1, value=label)
    if indent:
        cell.font = _INDENT_FONT
        cell.alignment = Alignment(indent=2)
    elif bold:
        cell.font = _BOLD_FONT
    if border:
        cell.border = border

    actual_style = style_name + "_bold" if bold else style_name
    for i, v in enumerate(values):
        c = ws.cell(row=row, column=2 + i, value=v)
        c.style = actual_style
        if border:
            c.border = border

    if yearly_totals:
        for j, yt in enumerate(yearly_totals):
            c = ws.cell(row=row, column=38 + j, value=yt)
            c.style = actual_style
            c.font = Font(bold=True, size=10)
            if border:
                c.border = border


def _sum_year(arr: list[float], year: int) -> float:
    s = (year - 1) * 12
    return sum(arr[s:s + 12])


def _yearly_totals(arr: list[float]) -> list[float]:
    return [_sum_year(arr, 1), _sum_year(arr, 2), _sum_year(arr, 3)]


# ---------------------------------------------------------------------------
# Tab 1: Assumptions
# ---------------------------------------------------------------------------

def _build_assumptions_tab(ws, config: dict):
    ws.title = "Assumptions"
    _set_col_widths(ws, {"A": 35, "B": 25, "C": 20, "D": 20})

    row = 1
    ws.cell(row=row, column=1, value=f"Tibetan Spirit Financial Model {config['meta']['version']} — Assumptions")
    ws.cell(row=row, column=1).font = _TITLE_FONT
    row += 2

    # Meta
    _write_section_header(ws, row, "Model Meta", 4)
    row += 1
    meta = config["meta"]
    for key in ["version", "scenario", "period", "months", "start_year", "start_month"]:
        ws.cell(row=row, column=1, value=key.replace("_", " ").title())
        ws.cell(row=row, column=1).font = _INDENT_FONT
        ws.cell(row=row, column=2, value=meta[key])
        row += 1

    row += 1
    # D2C
    _write_section_header(ws, row, "D2C Product Business", 4)
    row += 1
    d2c = config["d2c"]
    if "cogs_pct" in d2c:
        ws.cell(row=row, column=1, value="  Blended COGS %").font = _INDENT_FONT
        ws.cell(row=row, column=2, value=d2c["cogs_pct"])
        ws.cell(row=row, column=2).style = "pct"
        row += 1
    ws.cell(row=row, column=1, value="  Shipping/Fulfillment %").font = _INDENT_FONT
    ws.cell(row=row, column=2, value=d2c["shipping_fulfillment_pct"])
    ws.cell(row=row, column=2).style = "pct"
    row += 2

    # Product Lines / Channels
    lines_key = "product_lines" if "product_lines" in d2c else "channels"
    section_label = "Product Line Assumptions" if lines_key == "product_lines" else "Channel Assumptions"
    _write_section_header(ws, row, section_label, 4)
    row += 1
    _write_header_row(ws, row, ["Product Line", "Start Month", "Starting Rev", "Growth / Type"])
    row += 1
    for ln_name, ln_cfg in d2c[lines_key].items():
        ws.cell(row=row, column=1, value=ln_cfg.get("label", ln_name))
        ws.cell(row=row, column=1).font = _INDENT_FONT
        ws.cell(row=row, column=2, value=ln_cfg["start_month"])
        ln_type = ln_cfg.get("type", "standard")
        if ln_type == "quarterly":
            ws.cell(row=row, column=3, value=ln_cfg["starting_quarterly_revenue"])
            ws.cell(row=row, column=3).style = "dollar"
            ws.cell(row=row, column=4, value=f"Quarterly, {ln_cfg['quarterly_growth']:.0%} growth/qtr")
        elif ln_type == "unit_based":
            ws.cell(row=row, column=3, value=ln_cfg["avg_order_value"])
            ws.cell(row=row, column=3).style = "dollar"
            ws.cell(row=row, column=4, value=f"Unit-based, {ln_cfg['units_per_month']} units/mo")
        else:
            ws.cell(row=row, column=3, value=ln_cfg["starting_revenue"])
            ws.cell(row=row, column=3).style = "dollar"
            growth = ln_cfg["monthly_growth"]
            ws.cell(row=row, column=4, value=f"Monthly: {growth}")
        row += 1
        # COGS rate
        cogs = ln_cfg.get("cogs_pct", ln_cfg.get("cogs_pct_override"))
        fee = ln_cfg.get("platform_fee_pct")
        ws.cell(row=row, column=1, value="    COGS %")
        ws.cell(row=row, column=1).font = Font(color="A0A0A0", size=9)
        if cogs:
            ws.cell(row=row, column=2, value=cogs)
            ws.cell(row=row, column=2).style = "pct"
        if fee:
            ws.cell(row=row, column=3, value=fee)
            ws.cell(row=row, column=3).style = "pct"
        row += 1

    row += 1
    # Travels
    _write_section_header(ws, row, "TS Travels (Pilgrimages)", 4)
    row += 1
    trav = config["travels"]
    ws.cell(row=row, column=1, value="  Trips per year").font = _INDENT_FONT
    ws.cell(row=row, column=2, value=str(trav["trips_per_year"]))
    row += 1
    ws.cell(row=row, column=1, value="  Revenue per trip").font = _INDENT_FONT
    ws.cell(row=row, column=2, value=trav["revenue_per_trip"])
    ws.cell(row=row, column=2).style = "dollar"
    row += 1
    ws.cell(row=row, column=1, value="  COGS %").font = _INDENT_FONT
    ws.cell(row=row, column=2, value=trav["cogs_pct"])
    ws.cell(row=row, column=2).style = "pct"
    row += 2

    # Costs
    _write_section_header(ws, row, "Operating Costs", 4)
    row += 1
    costs = config["costs"]
    # Personnel
    ws.cell(row=row, column=1, value="Personnel").font = _BOLD_FONT
    row += 1
    for k, v in costs["personnel"].items():
        ws.cell(row=row, column=1, value=f"  {k}").font = _INDENT_FONT
        ws.cell(row=row, column=2, value=v)
        ws.cell(row=row, column=2).style = "dollar"
        row += 1
    # Technology
    ws.cell(row=row, column=1, value="Technology").font = _BOLD_FONT
    row += 1
    for k, v in costs["technology"].items():
        ws.cell(row=row, column=1, value=f"  {k}").font = _INDENT_FONT
        ws.cell(row=row, column=2, value=v)
        ws.cell(row=row, column=2).style = "dollar"
        row += 1
    # Marketing
    ws.cell(row=row, column=1, value="Marketing").font = _BOLD_FONT
    row += 1
    mkt = costs["marketing"]
    if "ramp_schedule" in mkt:
        ws.cell(row=row, column=1, value="  Y1 Ramp Schedule").font = _INDENT_FONT
        row += 1
        for entry in mkt["ramp_schedule"]:
            ws.cell(row=row, column=1,
                    value=f"    Through M{entry['through_month']}").font = Font(color="A0A0A0", size=9)
            ws.cell(row=row, column=2, value=entry["pct"])
            ws.cell(row=row, column=2).style = "pct"
            row += 1
        for k in ["pct_of_product_revenue_y2", "pct_of_product_revenue_y3"]:
            ws.cell(row=row, column=1, value=f"  {k}").font = _INDENT_FONT
            ws.cell(row=row, column=2, value=mkt[k])
            ws.cell(row=row, column=2).style = "pct"
            row += 1
        for yr in [1, 2, 3]:
            key = f"holiday_multiplier_y{yr}"
            if key in mkt:
                ws.cell(row=row, column=1, value=f"  Holiday Multiplier Y{yr}").font = _INDENT_FONT
                ws.cell(row=row, column=2, value=mkt[key])
                row += 1
    else:
        for k in ["pct_of_product_revenue_y1", "pct_of_product_revenue_y2",
                   "pct_of_product_revenue_y3"]:
            ws.cell(row=row, column=1, value=f"  {k}").font = _INDENT_FONT
            ws.cell(row=row, column=2, value=mkt[k])
            ws.cell(row=row, column=2).style = "pct"
            row += 1
        ws.cell(row=row, column=1, value="  Q4 Multiplier").font = _INDENT_FONT
        ws.cell(row=row, column=2, value=mkt["q4_multiplier"])
        row += 1
    # Seller payout
    ws.cell(row=row, column=1, value="Seller Payout").font = _BOLD_FONT
    row += 1
    ws.cell(row=row, column=1, value="  Monthly (Y1-5)").font = _INDENT_FONT
    ws.cell(row=row, column=2, value=costs["seller_payout"]["monthly_y1_5"])
    ws.cell(row=row, column=2).style = "dollar"
    row += 1
    # Foundation
    ws.cell(row=row, column=1, value="Foundation (Dharma Giving)").font = _BOLD_FONT
    row += 1
    ws.cell(row=row, column=1, value="  % of Net Profit").font = _INDENT_FONT
    ws.cell(row=row, column=2, value=costs["foundation"]["pct_of_net_profit"])
    ws.cell(row=row, column=2).style = "pct"
    row += 2

    # Capital
    _write_section_header(ws, row, "Capital Structure", 4)
    row += 1
    cap = config["capital"]
    ws.cell(row=row, column=1, value="  Cash on Hand").font = _INDENT_FONT
    ws.cell(row=row, column=2, value=cap["cash_on_hand"])
    ws.cell(row=row, column=2).style = "dollar"
    row += 1
    for yr in ["y1", "y2", "y3"]:
        ws.cell(row=row, column=1, value=f"  Ops Capital {yr.upper()}").font = _INDENT_FONT
        ws.cell(row=row, column=2, value=cap["ops_capital"][yr])
        ws.cell(row=row, column=2).style = "dollar"
        row += 1
    ws.cell(row=row, column=1, value="  Truist CD").font = _INDENT_FONT
    ws.cell(row=row, column=2, value=cap["truist_cd"])
    ws.cell(row=row, column=2).style = "dollar"


# ---------------------------------------------------------------------------
# Tab 2: D2C Monthly P&L
# ---------------------------------------------------------------------------

def _build_d2c_pnl_tab(ws, pnl: dict, config: dict):
    ws.title = "D2C Monthly P&L"
    lines_key = "product_lines" if "product_lines" in config["d2c"] else "channels"
    _set_col_widths(ws, {1: 30})
    # Set month columns to 10
    for c in range(2, 41):
        ws.column_dimensions[get_column_letter(c)].width = 11

    row = 1
    # Header row: labels + 36 months + Y1/Y2/Y3
    headers = [""] + pnl["month_labels"] + ["Y1 Total", "Y2 Total", "Y3 Total"]
    _write_header_row(ws, row, headers)
    row += 2

    ys = pnl["yearly_summary"]

    # --- Revenue section ---
    _write_section_header(ws, row, "REVENUE", 40)
    row += 1

    for ch_name, ch_vals in pnl["d2c_channels"].items():
        label = config["d2c"][lines_key][ch_name].get("label", ch_name)
        _write_monthly_row(ws, row, f"  {label}", ch_vals, indent=True,
                           yearly_totals=_yearly_totals(ch_vals))
        row += 1

    _write_monthly_row(ws, row, "Total D2C Revenue", pnl["d2c_total"],
                       bold=True, yearly_totals=[ys[y]["d2c_total"] for y in ["Y1", "Y2", "Y3"]],
                       border=_TOTAL_BORDER)
    row += 2

    # --- COGS section ---
    _write_section_header(ws, row, "COST OF GOODS SOLD", 40)
    row += 1

    for ch_name, ch_vals in pnl["cogs_by_channel"].items():
        label = config["d2c"][lines_key][ch_name].get("label", ch_name)
        _write_monthly_row(ws, row, f"  {label}", ch_vals, indent=True,
                           yearly_totals=_yearly_totals(ch_vals))
        row += 1

    # D2C-only COGS total (exclude travels)
    d2c_cogs = [sum(pnl["cogs_by_channel"][ch][i] for ch in pnl["cogs_by_channel"])
                for i in range(36)]
    _write_monthly_row(ws, row, "Total D2C COGS", d2c_cogs,
                       bold=True, yearly_totals=_yearly_totals(d2c_cogs),
                       border=_TOTAL_BORDER)
    row += 2

    # --- Gross Profit (D2C only) ---
    d2c_gp = [pnl["d2c_total"][i] - d2c_cogs[i] for i in range(36)]
    _write_section_header(ws, row, "GROSS PROFIT (D2C)", 40)
    row += 1
    _write_monthly_row(ws, row, "D2C Gross Profit", d2c_gp,
                       bold=True, yearly_totals=_yearly_totals(d2c_gp),
                       border=_TOTAL_BORDER)
    row += 1
    # Gross margin row
    d2c_gm = [d2c_gp[i] / pnl["d2c_total"][i] if pnl["d2c_total"][i] > 0 else 0
              for i in range(36)]
    yearly_gm = [_sum_year(d2c_gp, y) / _sum_year(pnl["d2c_total"], y)
                 if _sum_year(pnl["d2c_total"], y) > 0 else 0 for y in [1, 2, 3]]
    _write_monthly_row(ws, row, "  Gross Margin %", d2c_gm, style_name="pct",
                       indent=True, yearly_totals=yearly_gm)
    row += 2

    # --- Operating Expenses ---
    _write_section_header(ws, row, "OPERATING EXPENSES", 40)
    row += 1

    cost_lines = [
        ("Personnel", pnl["costs"]["personnel"]),
        ("Technology", pnl["costs"]["technology"]),
        ("Marketing", pnl["costs"]["marketing"]),
        ("Shipping & Fulfillment", pnl["costs"]["shipping"]),
        ("Platform Fees", pnl["costs"]["platform_fees"]),
        ("Seller Payout", pnl["costs"]["seller_payout"]),
        ("Foundation (Dharma Giving)", pnl["costs"]["foundation"]),
    ]
    for label, vals in cost_lines:
        _write_monthly_row(ws, row, f"  {label}", vals, indent=True,
                           yearly_totals=_yearly_totals(vals))
        row += 1

    _write_monthly_row(ws, row, "Total OpEx", pnl["total_opex"],
                       bold=True, yearly_totals=[ys[y]["total_opex"] for y in ["Y1", "Y2", "Y3"]],
                       border=_TOTAL_BORDER)
    row += 2

    # --- D2C EBITDA ---
    d2c_ebitda = [d2c_gp[i] - pnl["total_opex"][i] for i in range(36)]
    _write_section_header(ws, row, "D2C EBITDA", 40)
    row += 1
    _write_monthly_row(ws, row, "D2C EBITDA", d2c_ebitda,
                       bold=True, yearly_totals=_yearly_totals(d2c_ebitda),
                       border=_TOTAL_BORDER)

    # Freeze pane: col A + header row
    ws.freeze_panes = "B2"


# ---------------------------------------------------------------------------
# Tab 3: TS Travels
# ---------------------------------------------------------------------------

def _build_travels_tab(ws, pnl: dict):
    ws.title = "TS Travels"
    _set_col_widths(ws, {1: 30})
    for c in range(2, 41):
        ws.column_dimensions[get_column_letter(c)].width = 11

    row = 1
    headers = [""] + pnl["month_labels"] + ["Y1 Total", "Y2 Total", "Y3 Total"]
    _write_header_row(ws, row, headers)
    row += 2

    ys = pnl["yearly_summary"]

    _write_section_header(ws, row, "TRAVELS REVENUE", 40)
    row += 1
    _write_monthly_row(ws, row, "TS Travels Revenue", pnl["travels"],
                       bold=True, yearly_totals=[ys[y]["travels"] for y in ["Y1", "Y2", "Y3"]],
                       border=_THIN_BORDER)
    row += 2

    _write_section_header(ws, row, "TRAVELS COGS (60%)", 40)
    row += 1
    _write_monthly_row(ws, row, "Travels COGS", pnl["cogs_travels"],
                       bold=True, yearly_totals=_yearly_totals(pnl["cogs_travels"]),
                       border=_THIN_BORDER)
    row += 2

    _write_section_header(ws, row, "TRAVELS GROSS PROFIT", 40)
    row += 1
    trav_gp = [pnl["travels"][i] - pnl["cogs_travels"][i] for i in range(36)]
    _write_monthly_row(ws, row, "Travels Gross Profit", trav_gp,
                       bold=True, yearly_totals=_yearly_totals(trav_gp),
                       border=_TOTAL_BORDER)
    row += 1
    trav_gm = [trav_gp[i] / pnl["travels"][i] if pnl["travels"][i] > 0 else 0
               for i in range(36)]
    yearly_trav_gm = [_sum_year(trav_gp, y) / _sum_year(pnl["travels"], y)
                      if _sum_year(pnl["travels"], y) > 0 else 0 for y in [1, 2, 3]]
    _write_monthly_row(ws, row, "  Gross Margin %", trav_gm, style_name="pct",
                       indent=True, yearly_totals=yearly_trav_gm)

    ws.freeze_panes = "B2"


# ---------------------------------------------------------------------------
# Tab 4: Combined P&L
# ---------------------------------------------------------------------------

def _build_combined_pnl_tab(ws, pnl: dict):
    ws.title = "Combined P&L"
    _set_col_widths(ws, {1: 30})
    for c in range(2, 41):
        ws.column_dimensions[get_column_letter(c)].width = 11

    row = 1
    headers = [""] + pnl["month_labels"] + ["Y1 Total", "Y2 Total", "Y3 Total"]
    _write_header_row(ws, row, headers)
    row += 2

    ys = pnl["yearly_summary"]

    # Revenue
    _write_section_header(ws, row, "REVENUE", 40)
    row += 1
    _write_monthly_row(ws, row, "  D2C Revenue", pnl["d2c_total"], indent=True,
                       yearly_totals=[ys[y]["d2c_total"] for y in ["Y1", "Y2", "Y3"]])
    row += 1
    _write_monthly_row(ws, row, "  TS Travels", pnl["travels"], indent=True,
                       yearly_totals=[ys[y]["travels"] for y in ["Y1", "Y2", "Y3"]])
    row += 1
    _write_monthly_row(ws, row, "Total Revenue", pnl["total_revenue"],
                       bold=True,
                       yearly_totals=[ys[y]["total_revenue"] for y in ["Y1", "Y2", "Y3"]],
                       border=_TOTAL_BORDER)
    row += 2

    # COGS
    _write_section_header(ws, row, "COST OF GOODS SOLD", 40)
    row += 1
    d2c_cogs = [pnl["total_cogs"][i] - pnl["cogs_travels"][i] for i in range(36)]
    _write_monthly_row(ws, row, "  D2C COGS", d2c_cogs, indent=True,
                       yearly_totals=_yearly_totals(d2c_cogs))
    row += 1
    _write_monthly_row(ws, row, "  Travels COGS", pnl["cogs_travels"], indent=True,
                       yearly_totals=_yearly_totals(pnl["cogs_travels"]))
    row += 1
    _write_monthly_row(ws, row, "Total COGS", pnl["total_cogs"],
                       bold=True,
                       yearly_totals=[ys[y]["total_cogs"] for y in ["Y1", "Y2", "Y3"]],
                       border=_TOTAL_BORDER)
    row += 2

    # Gross Profit
    _write_section_header(ws, row, "GROSS PROFIT", 40)
    row += 1
    _write_monthly_row(ws, row, "Gross Profit", pnl["gross_profit"],
                       bold=True,
                       yearly_totals=[ys[y]["gross_profit"] for y in ["Y1", "Y2", "Y3"]],
                       border=_TOTAL_BORDER)
    row += 1
    gm = [pnl["gross_profit"][i] / pnl["total_revenue"][i]
          if pnl["total_revenue"][i] > 0 else 0 for i in range(36)]
    yearly_gm = [ys[y]["gross_profit"] / ys[y]["total_revenue"]
                 if ys[y]["total_revenue"] > 0 else 0 for y in ["Y1", "Y2", "Y3"]]
    _write_monthly_row(ws, row, "  Gross Margin %", gm, style_name="pct",
                       indent=True, yearly_totals=yearly_gm)
    row += 2

    # OpEx
    _write_section_header(ws, row, "OPERATING EXPENSES", 40)
    row += 1
    cost_lines = [
        ("Personnel", pnl["costs"]["personnel"]),
        ("Technology", pnl["costs"]["technology"]),
        ("Marketing", pnl["costs"]["marketing"]),
        ("Shipping & Fulfillment", pnl["costs"]["shipping"]),
        ("Platform Fees", pnl["costs"]["platform_fees"]),
        ("Seller Payout", pnl["costs"]["seller_payout"]),
        ("Foundation (Dharma Giving)", pnl["costs"]["foundation"]),
    ]
    for label, vals in cost_lines:
        _write_monthly_row(ws, row, f"  {label}", vals, indent=True,
                           yearly_totals=_yearly_totals(vals))
        row += 1
    _write_monthly_row(ws, row, "Total OpEx", pnl["total_opex"],
                       bold=True,
                       yearly_totals=[ys[y]["total_opex"] for y in ["Y1", "Y2", "Y3"]],
                       border=_TOTAL_BORDER)
    row += 2

    # EBITDA
    _write_section_header(ws, row, "EBITDA", 40)
    row += 1
    _write_monthly_row(ws, row, "EBITDA", pnl["ebitda"],
                       bold=True,
                       yearly_totals=[ys[y]["ebitda"] for y in ["Y1", "Y2", "Y3"]],
                       border=_TOTAL_BORDER)
    row += 2

    # Cash Flow
    _write_section_header(ws, row, "CASH FLOW", 40)
    row += 1
    _write_monthly_row(ws, row, "  Capital Infusion", pnl["capital_infusion"],
                       indent=True, yearly_totals=_yearly_totals(pnl["capital_infusion"]))
    row += 1
    _write_monthly_row(ws, row, "Ending Cash", pnl["ending_cash"],
                       bold=True,
                       yearly_totals=[pnl["ending_cash"][11],
                                      pnl["ending_cash"][23],
                                      pnl["ending_cash"][35]],
                       border=_TOTAL_BORDER)

    ws.freeze_panes = "B2"


# ---------------------------------------------------------------------------
# Tab 5: Channel Scenarios
# ---------------------------------------------------------------------------

def _build_scenarios_tab(ws, scenarios: dict):
    ws.title = "Channel Scenarios"
    _set_col_widths(ws, {1: 25})

    row = 1
    ws.cell(row=row, column=1, value="Channel Scenario Comparison")
    ws.cell(row=row, column=1).font = _TITLE_FONT
    row += 2

    # Build header: Metric | Scenario1 Y1 Y2 Y3 | Scenario2 Y1 Y2 Y3 ...
    scenario_names = list(scenarios.keys())
    header = ["Metric"]
    for name in scenario_names:
        header += [f"{name} Y1", f"{name} Y2", f"{name} Y3"]
    _write_header_row(ws, row, header)

    # Set widths for scenario columns
    for c in range(2, 2 + len(scenario_names) * 3):
        ws.column_dimensions[get_column_letter(c)].width = 16
    row += 1

    metrics = [
        ("Total Revenue", "total_revenue"),
        ("D2C Revenue", "d2c_total"),
        ("Travels Revenue", "travels"),
        ("Gross Profit", "gross_profit"),
        ("Marketing Spend", "marketing"),
        ("Total OpEx", "total_opex"),
        ("EBITDA", "ebitda"),
    ]

    for label, key in metrics:
        is_bold = label in ("Total Revenue", "Gross Profit", "EBITDA")
        cell = ws.cell(row=row, column=1, value=label)
        cell.font = _BOLD_FONT if is_bold else Font(size=10)
        col = 2
        for name in scenario_names:
            s_pnl = scenarios[name]
            ys = s_pnl["yearly_summary"]
            for yr in ["Y1", "Y2", "Y3"]:
                c = ws.cell(row=row, column=col, value=ys[yr][key])
                c.style = "dollar_bold" if is_bold else "dollar"
                col += 1
        if is_bold:
            for c_idx in range(1, col):
                ws.cell(row=row, column=c_idx).border = _THIN_BORDER
        row += 1

    # Ending Cash — special handling (point-in-time, not from yearly_summary)
    row_label = "Ending Cash"
    cell = ws.cell(row=row, column=1, value=row_label)
    cell.font = _BOLD_FONT
    col = 2
    for name in scenario_names:
        s_pnl = scenarios[name]
        for idx in [11, 23, 35]:
            c = ws.cell(row=row, column=col, value=s_pnl["ending_cash"][idx])
            c.style = "dollar_bold"
            col += 1
    for c_idx in range(1, col):
        ws.cell(row=row, column=c_idx).border = _TOTAL_BORDER


# ---------------------------------------------------------------------------
# Tab 6: Sensitivity
# ---------------------------------------------------------------------------

def _build_sensitivity_tab(ws, pnl: dict, config: dict):
    ws.title = "Sensitivity"
    _set_col_widths(ws, {1: 25, 2: 15, 3: 15, 4: 15, 5: 15, 6: 15, 7: 15, 8: 15})

    row = 1
    ws.cell(row=row, column=1, value="Revenue Sensitivity Analysis")
    ws.cell(row=row, column=1).font = _TITLE_FONT
    row += 2

    sens = config.get("sensitivity", {})
    factors = sens.get("revenue_factors", [0.70, 0.85, 1.0, 1.15])
    labels = sens.get("labels", [f"{f:.0%}" for f in factors])

    ys = pnl["yearly_summary"]
    base_rev = [ys[y]["total_revenue"] for y in ["Y1", "Y2", "Y3"]]
    base_ebitda = [ys[y]["ebitda"] for y in ["Y1", "Y2", "Y3"]]
    # Fixed costs (approx): personnel + technology + seller_payout
    fixed_costs = [
        ys[y]["personnel"] + ys[y]["technology"] + ys[y]["seller_payout"]
        for y in ["Y1", "Y2", "Y3"]
    ]

    # Revenue table
    _write_section_header(ws, row, "Total Revenue by Scenario", 8)
    row += 1
    _write_header_row(ws, row, ["Scenario", "Y1 Revenue", "Y2 Revenue", "Y3 Revenue"])
    row += 1
    for i, factor in enumerate(factors):
        is_base = abs(factor - 1.0) < 0.001
        cell = ws.cell(row=row, column=1, value=labels[i])
        cell.font = _BOLD_FONT if is_base else Font(size=10)
        for j in range(3):
            c = ws.cell(row=row, column=2 + j, value=base_rev[j] * factor)
            c.style = "dollar_bold" if is_base else "dollar"
        row += 1

    row += 1
    # EBITDA table — scale variable costs proportionally, keep fixed costs fixed
    _write_section_header(ws, row, "EBITDA by Scenario (directional)", 8)
    row += 1
    _write_header_row(ws, row, ["Scenario", "Y1 EBITDA", "Y2 EBITDA", "Y3 EBITDA"])
    row += 1
    for i, factor in enumerate(factors):
        is_base = abs(factor - 1.0) < 0.001
        cell = ws.cell(row=row, column=1, value=labels[i])
        cell.font = _BOLD_FONT if is_base else Font(size=10)
        for j in range(3):
            # EBITDA = base_ebitda + (factor-1) * (base_rev - fixed_costs)
            # This approximates: variable profit scales with revenue, fixed stays fixed
            variable_profit = base_rev[j] - fixed_costs[j] - (base_rev[j] - base_ebitda[j] - fixed_costs[j])
            # Simpler: ebitda_scaled = fixed_costs stay, everything else scales
            # ebitda = revenue*factor - cogs*factor - variable_opex*factor - fixed_opex
            # variable_opex = total_opex - fixed_costs
            total_opex_yr = ys[["Y1", "Y2", "Y3"][j]]["total_opex"]
            total_cogs_yr = ys[["Y1", "Y2", "Y3"][j]]["total_cogs"]
            variable_opex = total_opex_yr - fixed_costs[j]
            scaled_ebitda = (base_rev[j] * factor
                             - total_cogs_yr * factor
                             - variable_opex * factor
                             - fixed_costs[j])
            c = ws.cell(row=row, column=2 + j, value=scaled_ebitda)
            c.style = "dollar_bold" if is_base else "dollar"
        row += 1

    row += 1
    # Note
    ws.cell(row=row, column=1,
            value="Note: EBITDA sensitivity is directional. Fixed costs (personnel, "
                  "technology, seller payout) held constant; all other costs scale "
                  "proportionally with revenue.")
    ws.cell(row=row, column=1).font = Font(italic=True, color="808080", size=9)


# ---------------------------------------------------------------------------
# Tab 7: Capital & Cash
# ---------------------------------------------------------------------------

def _build_capital_tab(ws, pnl: dict, config: dict):
    ws.title = "Capital & Cash"
    _set_col_widths(ws, {1: 35, 2: 18, 3: 18, 4: 18, 5: 18})

    row = 1
    ws.cell(row=row, column=1, value="Capital Deployment & Cash Position")
    ws.cell(row=row, column=1).font = _TITLE_FONT
    row += 2

    cap = config["capital"]

    # Capital sources
    _write_section_header(ws, row, "Capital Sources", 5)
    row += 1
    _write_header_row(ws, row, ["Source", "Amount", "Timing"])
    row += 1

    sources = [
        ("Cash on Hand", cap["cash_on_hand"], "At model start"),
        ("Ops Capital Y1", cap["ops_capital"]["y1"], "Month 1 (Jun 2026)"),
        ("Ops Capital Y2", cap["ops_capital"]["y2"], "Month 13 (Jun 2027)"),
        ("Ops Capital Y3", cap["ops_capital"]["y3"], "Month 25 (Jun 2028)"),
        ("Truist CD", cap["truist_cd"], "Month 10 (Mar 2027)"),
    ]
    total_capital = sum(s[1] for s in sources)
    for label, amt, timing in sources:
        ws.cell(row=row, column=1, value=f"  {label}").font = _INDENT_FONT
        ws.cell(row=row, column=2, value=amt).style = "dollar"
        ws.cell(row=row, column=3, value=timing)
        row += 1
    ws.cell(row=row, column=1, value="Total Capital").font = _BOLD_FONT
    ws.cell(row=row, column=2, value=total_capital).style = "dollar_bold"
    for c in range(1, 4):
        ws.cell(row=row, column=c).border = _TOTAL_BORDER
    row += 2

    # Inventory facility
    _write_section_header(ws, row, "Inventory Facility", 5)
    row += 1
    inv = cap["inventory_facility"]
    ws.cell(row=row, column=1, value="  Cap").font = _INDENT_FONT
    ws.cell(row=row, column=2, value=inv["cap"]).style = "dollar"
    row += 1
    ws.cell(row=row, column=1, value="  Interest Rate").font = _INDENT_FONT
    ws.cell(row=row, column=2, value=inv["interest"]).style = "pct"
    row += 1
    ws.cell(row=row, column=1, value="  Reconciliation").font = _INDENT_FONT
    ws.cell(row=row, column=2, value=inv["reconciliation"])
    row += 2

    # Seller payout schedule
    _write_section_header(ws, row, "Seller Payout Schedule", 5)
    row += 1
    sp = config["costs"]["seller_payout"]
    ws.cell(row=row, column=1, value="  Monthly (Y1-5)").font = _INDENT_FONT
    ws.cell(row=row, column=2, value=sp["monthly_y1_5"]).style = "dollar"
    ws.cell(row=row, column=3, value="$36K/year guaranteed")
    row += 1
    ws.cell(row=row, column=1, value="  Monthly (Y6-10)").font = _INDENT_FONT
    ws.cell(row=row, column=2, value=sp["monthly_y6_10"]).style = "dollar"
    ws.cell(row=row, column=3, value="Reduced — monk ordination")
    row += 1
    ws.cell(row=row, column=1, value="  36-Month Total").font = _BOLD_FONT
    ws.cell(row=row, column=2, value=sp["monthly_y1_5"] * 36).style = "dollar_bold"
    for c in range(1, 4):
        ws.cell(row=row, column=c).border = _TOTAL_BORDER
    row += 2

    # Year-end cash positions
    _write_section_header(ws, row, "Year-End Cash Positions", 5)
    row += 1
    _write_header_row(ws, row, ["Period", "Ending Cash", "Cumulative EBITDA",
                                "Capital Deployed"])
    row += 1

    ys = pnl["yearly_summary"]
    cum_ebitda = 0.0
    cum_capital = 0.0
    for yr_idx, yr_label in enumerate(["Y1", "Y2", "Y3"]):
        month_idx = (yr_idx + 1) * 12 - 1  # 11, 23, 35
        cum_ebitda += ys[yr_label]["ebitda"]
        cum_capital += _sum_year(pnl["capital_infusion"], yr_idx + 1)

        ws.cell(row=row, column=1, value=f"  End of {yr_label}").font = _INDENT_FONT
        ws.cell(row=row, column=2, value=pnl["ending_cash"][month_idx]).style = "dollar"
        ws.cell(row=row, column=3, value=cum_ebitda).style = "dollar"
        ws.cell(row=row, column=4, value=cum_capital).style = "dollar"
        row += 1

    # Cash on hand at start for reference
    row += 1
    ws.cell(row=row, column=1, value="Starting Cash (model start)").font = _BOLD_FONT
    ws.cell(row=row, column=2, value=cap["cash_on_hand"]).style = "dollar_bold"


# ---------------------------------------------------------------------------
# Main builder
# ---------------------------------------------------------------------------

def build_workbook(config: dict, output_path: Path) -> dict:
    """Build the full workbook and return summary metrics."""
    pnl = build_monthly_pnl(config)
    scenarios = build_scenarios(config)

    wb = Workbook()
    _make_styles(wb)

    # Tab 1: Assumptions
    ws1 = wb.active
    _build_assumptions_tab(ws1, config)

    # Tab 2: D2C Monthly P&L
    ws2 = wb.create_sheet()
    _build_d2c_pnl_tab(ws2, pnl, config)

    # Tab 3: TS Travels
    ws3 = wb.create_sheet()
    _build_travels_tab(ws3, pnl)

    # Tab 4: Combined P&L
    ws4 = wb.create_sheet()
    _build_combined_pnl_tab(ws4, pnl)

    # Tab 5: Channel Scenarios
    ws5 = wb.create_sheet()
    _build_scenarios_tab(ws5, scenarios)

    # Tab 6: Sensitivity
    ws6 = wb.create_sheet()
    _build_sensitivity_tab(ws6, pnl, config)

    # Tab 7: Capital & Cash
    ws7 = wb.create_sheet()
    _build_capital_tab(ws7, pnl, config)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))

    ys = pnl["yearly_summary"]
    return {
        "y1_revenue": ys["Y1"]["total_revenue"],
        "y2_revenue": ys["Y2"]["total_revenue"],
        "y3_revenue": ys["Y3"]["total_revenue"],
        "y1_ebitda": ys["Y1"]["ebitda"],
        "y2_ebitda": ys["Y2"]["ebitda"],
        "y3_ebitda": ys["Y3"]["ebitda"],
        "y1_ending_cash": pnl["ending_cash"][11],
        "y2_ending_cash": pnl["ending_cash"][23],
        "y3_ending_cash": pnl["ending_cash"][35],
    }


def main():
    parser = argparse.ArgumentParser(
        description="Build Tibetan Spirit v7 financial model xlsx"
    )
    parser.add_argument(
        "--config",
        default=str(Path(__file__).parent / "config" / "model_v8.yaml"),
        help="Path to YAML config file",
    )
    parser.add_argument(
        "--output",
        default=str(_PROJECT_ROOT / "deliverables" / "outputs" / "docs"
                     / "ts-financial-model-v8-2026-04-17.xlsx"),
        help="Path for output xlsx file",
    )
    args = parser.parse_args()

    config_path = Path(args.config)
    if not config_path.exists():
        print(f"ERROR: Config not found: {config_path}")
        sys.exit(1)

    with open(config_path) as f:
        config = yaml.safe_load(f)

    output_path = Path(args.output)
    summary = build_workbook(config, output_path)

    print(f"Workbook saved: {output_path}")
    print()
    print("=== 3-Year Summary ===")
    for yr in ["y1", "y2", "y3"]:
        label = yr.upper()
        rev = summary[f"{yr}_revenue"]
        ebitda = summary[f"{yr}_ebitda"]
        cash = summary[f"{yr}_ending_cash"]
        print(f"  {label}: Revenue ${rev:,.0f} | EBITDA ${ebitda:,.0f} | "
              f"Ending Cash ${cash:,.0f}")


if __name__ == "__main__":
    main()
