#!/usr/bin/env python3
"""
Extract v6 Excel financial model to structured JSON.

Reads TS-Financial-Model-v6-Moderate-Upside.xlsx and outputs a single JSON
file that serves as the canonical data source for deal docs and pitch deck.

Usage:
    python3 scripts/financial_model/extract_v6.py
"""
import json
import sys
from datetime import date
from pathlib import Path

import openpyxl

EXCEL_PATH = Path(
    "/Users/chrismauze/Documents/Claude/Projects"
    "/💼 TS / Norbu/TS-Financial-Model-v6-Moderate-Upside.xlsx"
)
REPO_ROOT = Path(__file__).resolve().parent.parent.parent
OUTPUT_DIR = REPO_ROOT / "data" / "financial-model"
OUTPUT_FILE = OUTPUT_DIR / f"v6-extract-{date.today().isoformat()}.json"


def cell(ws, row, col):
    """Return cell value, converting None to 0 for numeric fields."""
    v = ws.cell(row=row, column=col).value
    return v


def num(val, default=0):
    """Coerce value to float, falling back to default."""
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def extract_assumptions(wb):
    """Extract key-value pairs from the Assumptions tab."""
    ws = wb["Assumptions"]
    assumptions = {}
    for row in range(1, ws.max_row + 1):
        key = ws.cell(row=row, column=1).value
        val = ws.cell(row=row, column=2).value
        note = ws.cell(row=row, column=3).value
        if key and val is not None:
            entry = {"value": val}
            if note:
                entry["note"] = str(note)
            assumptions[str(key).strip()] = entry
    return assumptions


def extract_cogs_by_channel(wb):
    """Extract Product COGS by channel from the Product COGS tab."""
    ws = wb["Product COGS"]
    channels = {}
    for row in range(4, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        if not name:
            continue
        channels[str(name).strip()] = {
            "cogs_pct": num(ws.cell(row=row, column=2).value),
            "gross_margin": num(ws.cell(row=row, column=3).value),
            "notes": str(ws.cell(row=row, column=4).value or ""),
        }
    return channels


def extract_sensitivity(wb):
    """Extract scenario rows from the Sensitivity tab."""
    ws = wb["Sensitivity"]
    scenarios = []
    for row in range(4, ws.max_row + 1):
        name = ws.cell(row=row, column=1).value
        if not name:
            continue
        entry = {
            "scenario": str(name).strip(),
            "revenue_factor": str(ws.cell(row=row, column=2).value or ""),
            "y1_ebitda": num(ws.cell(row=row, column=3).value),
            "y2_ebitda": num(ws.cell(row=row, column=4).value),
            "y3_ebitda": num(ws.cell(row=row, column=5).value),
        }
        m36_cash = ws.cell(row=row, column=6).value
        if m36_cash is not None:
            entry["m36_cash"] = num(m36_cash)
        scenarios.append(entry)
    return scenarios


def extract_capital_summary(wb):
    """Extract capital deployment items from Capital Summary tab."""
    ws = wb["Capital Summary"]
    items = {}
    for row in range(1, ws.max_row + 1):
        key = ws.cell(row=row, column=1).value
        val = ws.cell(row=row, column=2).value
        if key:
            items[str(key).strip()] = val
    return items


def extract_monthly_pnl(wb):
    """Extract full monthly P&L — all rows, all months.

    Returns a dict keyed by row label (preserving leading spaces for
    indented items). Each value is a dict with Y1/Y2/Y3 totals plus
    individual month values keyed by the column header (e.g. "Jun 26").
    """
    ws = wb["Monthly P&L"]

    # Build column header map: col_index -> header string
    headers = {}
    for col in range(2, ws.max_column + 1):
        h = ws.cell(row=4, column=col).value
        if h:
            headers[col] = str(h)

    pnl = {}
    for row in range(6, ws.max_row + 1):
        label = ws.cell(row=row, column=1).value
        if not label:
            continue
        label_str = str(label)
        row_data = {}
        for col, header in headers.items():
            v = ws.cell(row=row, column=col).value
            if v is not None:
                row_data[header] = v
        if row_data:
            pnl[label_str] = row_data

    return pnl


def extract_summary_and_channels(pnl):
    """Derive summary and channel_revenue from the monthly P&L data."""
    years = ["Y1 Total", "Y2 Total", "Y3 Total"]

    def get_yearly(row_label):
        row = pnl.get(row_label, {})
        return {y: num(row.get(y)) for y in years}

    summary = {
        "revenue": get_yearly("TOTAL REVENUE"),
        "ebitda": get_yearly("EBITDA"),
        "gross_profit": get_yearly("GROSS PROFIT"),
        "product_revenue": get_yearly("  Product Revenue"),
        "travel_revenue": get_yearly("  Travel Revenue"),
        "marketing_spend": get_yearly("  Marketing & Advertising"),
        "ending_cash": get_yearly("ENDING CASH BALANCE"),
    }

    channel_keys = {
        "shopify_d2c": "  Shopify D2C",
        "amazon": "  Amazon",
        "etsy": "  Etsy",
        "wholesale": "  Wholesale",
        "quarterly_subscription": "  Quarterly Subscription (net)",
        "high_value_dropship": "  High-Value Dropship",
        "ts_travels": "  TS Travels (per-trip)",
    }

    channel_revenue = {}
    for clean_name, pnl_label in channel_keys.items():
        channel_revenue[clean_name] = get_yearly(pnl_label)

    return summary, channel_revenue


def main():
    if not EXCEL_PATH.exists():
        print(f"ERROR: Excel file not found at {EXCEL_PATH}")
        sys.exit(1)

    print(f"Reading {EXCEL_PATH.name} ...")
    wb = openpyxl.load_workbook(str(EXCEL_PATH), data_only=True)

    # Extract all sections
    pnl = extract_monthly_pnl(wb)
    summary, channel_revenue = extract_summary_and_channels(pnl)
    assumptions = extract_assumptions(wb)
    cogs = extract_cogs_by_channel(wb)
    sensitivity = extract_sensitivity(wb)
    capital = extract_capital_summary(wb)

    output = {
        "meta": {
            "source": EXCEL_PATH.name,
            "extracted": date.today().isoformat(),
            "model_version": "v6",
            "scenario": "Moderate Upside",
            "period": "Jun 2026 – May 2029 (36 months)",
        },
        "summary": summary,
        "channel_revenue": channel_revenue,
        "assumptions": assumptions,
        "cogs_by_channel": cogs,
        "sensitivity": sensitivity,
        "capital_summary": capital,
        "monthly_pnl": pnl,
    }

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_FILE, "w") as f:
        json.dump(output, f, indent=2, default=str)

    print(f"\nJSON saved to {OUTPUT_FILE}")
    print(f"  Size: {OUTPUT_FILE.stat().st_size:,} bytes")

    # Print verification summary
    print("\n--- Verification Summary ---")
    for label, key in [("Revenue", "revenue"), ("EBITDA", "ebitda")]:
        vals = summary[key]
        print(f"  {label}:")
        for y in ["Y1 Total", "Y2 Total", "Y3 Total"]:
            print(f"    {y}: ${vals[y]:,.0f}")

    print(f"\n  Shopify D2C Y1: ${channel_revenue['shopify_d2c']['Y1 Total']:,.0f}")
    print(f"  M36 Ending Cash: ${summary['ending_cash']['Y3 Total']:,.0f}")


if __name__ == "__main__":
    main()
